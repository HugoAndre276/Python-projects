# -*- coding: utf-8 -*-
#Importation de modules
import json
import csv
import wcwidth
import ftfy
import datetime
import time
from openpyxl import Workbook


workbook=Workbook()

#Définitions de classes
class Personne:
    def __init__(self):

        #Nom
        self.name=""

        #Liste des messages
        self.messages=[]

        #Liste des temps
        self.time=[]

        #Liste des mots utilisés
        self.mots=[]

        #Mots utilisés et leurs occurences
        self.dico={}

        #Listes des mots utilisés et de leurs occurences
        self.arrayMots=[]
        self.arrayOcc=[]

        #Listes des occurences de messages par heure (index i: intervalle horaire [i,i+1])
        self.heures=[]
        self.jours=[]

        #Dico des dates (année, numéro de semaine)
        self.dateDico={}
        self.dateListe=[]
        self.dateOcc=[]

        #Nombre de photos
        self.pic=0

        #Nombre de liens
        self.lien=0

        #Page de excel
        self.sheet=workbook.create_sheet()


#Définitions de fonctions

def getDate(string):
    """Entrée: date sous forme yyyy-mm-dd hh:mm:ss.xxxx
       Sortie: la date en timestamp"""
    d=datetime.datetime(int(string[:4]),int(string[5:7]),int(string[8:10]),int(string[11:13]),int(string[14:16]),int(string[17:19]))
    unixtime = time.mktime(d.timetuple())
    return(unixtime)

def mots(string,L):
    """
    Entrée: string: ligne de messages
            L: liste de mots
    Fonctionnement: ajoute à L les mots trouvés dans string
    """
    i=0
    space=[' ',',','.','?','!','\n','"',"'",'(',')','*',':']
    #pluriel=['s','x']
    while i < len(string):
        c=''
        while i < len(string) and string[i] not in space:
            c+=string[i]
            i+=1
        i+=1
        if c!='': #and c[-1] in pluriel:
            #c=c[:-1]
            L.append(c.lower())


def compter_dico(liste,dico):
    """Entrée: liste: liste de mots
               dico: dictionnaire d'occurences indexées par les mots
       Fonctionnement: si un mot de la liste est dans le dictionnaire,
                       on ajoute 1 à ses occurences sinon on l'ajoute
    """
    for mot in liste:
        if mot in dico:
            dico[mot]+=1
        else:
            dico[mot]=1

def dico_to_array(dico,arrayMots,arrayOcc):
    """Entrée: dico: dictionnaire de mots et d'occurences
               arrayMots: liste des mots
               arrayOccHugo: liste des occurences
       Fonctionnement: ajoute chaque mot de dico et son index dans la liste
    """
    for mot in dico:
        arrayMots.append(str(mot))
        arrayOcc.append(dico[mot])

def temps(time,heureList,jourListe,dateDico):
    """Entrée: time: liste des temps des messages (en ms)
               heureList: liste des occurences de messages par heure
                          (index i, heure h --> h entre i et i+1)
               jourListe: idem mais avec des jours de la semaine
               dateDico: dictionnaire indexé par date en year, weekNb qui contient les occurences de messages de cette semaine
       Fonctionnement: pour chaque temps de la liste, check si changement d'heure
                       si oui, on enlève 2h, sinon 1h et on augmente l'occurence
                       à l'index i dans la liste

    """
    for t in time:
        date=datetime.datetime.fromtimestamp(t/1000.0)
        if date.month<10 and date.month>3:
            t-=2*3600*1000
        elif date.month==10 and  date.day<27:
            t-=2*3600*1000
        elif date.month==10 and date.day==27 and date.hour<3:
            t-=2*3600*1000
        elif date.month==3 and date.day>3:
            t-=2*3600*1000
        elif date.month==3 and date.day==3 and date.hour>2:
            t-=2*3600*1000
        else:
            t-=3600*1000
        date=datetime.datetime.fromtimestamp(t/1000.0)
        heureList[date.hour]+=1
        jourListe[date.weekday()]+=1
        date_=date.isocalendar()[0],date.isocalendar()[1]
        if date_ in dateDico:
            dateDico[date_]+=1
        else:
            dateDico[date_]=1

days=["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]

Hugo = Personne()
Manon = Personne()

Hugo.name="Hugo"
Manon.name="Manon"

personnes=[Hugo,Manon]

#lire le json de Messenger
for f in ["message_1.json","message_2.json","message_3.json"]:
    with open(f) as file:
        data = json.load(file)
        for p in data['messages']:
            if 'content' in p and p['type']=='Generic':
                if p['sender_name']=="Manon Dsc":
                    Manon.messages.append(ftfy.fix_text(p['content']))
                    Manon.time.append(p['timestamp_ms'])
                else:
                    Hugo.messages.append(ftfy.fix_text(p['content']))
                    Hugo.time.append(p['timestamp_ms'])
            if 'photos' in p:
                if p['sender_name']=='Manon Dsc':
                    Manon.pic+=1
                else:
                    Hugo.pic+=1
            if p['type']=='Share':
                if p['sender_name']=='Manon Dsc':
                    Manon.lien+=1
                else:
                    Hugo.lien+=1
            if 'reactions' in p:
                if p['reactions'][0]['actor']=='Manon Dsc':
                    Manon.messages.append(ftfy.fix_text(p['reactions'][0]['reaction']))
                else:
                    Hugo.messages.append(ftfy.fix_text(p['reactions'][0]['reaction']))

#lire le CSV de Discord de Hugo
with open('messages_disc_hugo.csv',newline='',encoding='utf-8') as file:
    reader=csv.reader(file)
    for row in reader:
        Hugo.messages.append(ftfy.fix_text(row[2]))
        Hugo.time.append(getDate(row[1])*1000)

#lire le CSV de Discord de Manon
with open('messages_disc_manon.csv',newline='',encoding='utf-8') as file:
    reader=csv.reader(file)
    for row in reader:
        Manon.messages.append(ftfy.fix_text(row[2]))
        Manon.time.append(getDate(row[1])*1000)

for p in personnes:
    for i in range(7):
        p.jours.append(0)
    for i in range(24):
        p.heures.append(0)
    for x in p.messages:
        mots(x,p.mots)
    compter_dico(p.mots,p.dico)
    dico_to_array(p.dico,p.arrayMots,p.arrayOcc)
    temps(p.time,p.heures,p.jours,p.dateDico)
    dico_to_array(p.dateDico,p.dateListe,p.dateOcc)

    p.sheet["A1"]=p.name
    p.sheet["A2"]="Mots"
    for i in range(len(p.arrayMots)):
        p.sheet.cell(row=i+3,column=1).value=p.arrayMots[i]
        p.sheet.cell(row=i+3,column=2).value=p.arrayOcc[i]

    p.sheet["D2"]="Heures des messages"
    for i in range(24):
        p.sheet.cell(row=i+3,column=4).value=str(i)+"h-"+str(i+1)+"h"
        p.sheet.cell(row=i+3,column=5).value=p.heures[i]

    p.sheet["G2"]="Jours des messages"
    for i in range(7):
        p.sheet.cell(row=i+3,column=7).value=days[i]
        p.sheet.cell(row=i+3,column=8).value=p.jours[i]

    p.sheet["J2"]="Semaines des messages"
    for i in range(1,len(p.dateListe)):
        p.sheet.cell(row=i+3,column=10).value=p.dateListe[len(p.dateListe)-i]
        p.sheet.cell(row=i+3,column=11).value=p.dateOcc[len(p.dateListe)-i]

    p.sheet["M2"]="Nombre de photos"
    p.sheet["M3"]=p.pic

    p.sheet["O2"]="Nombre de liens"
    p.sheet["O3"]=p.lien

del workbook["Sheet"]
workbook.save(filename="Manon.xlsx")
